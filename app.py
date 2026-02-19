import os
import sys
import io
import base64
import webbrowser
from threading import Timer
from datetime import datetime

from flask import Flask, render_template, request, url_for, redirect, flash
import qrcode
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# --- AUTH IMPORTS ---
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt

# ===============================
# PATH HANDLING (for EXE support)
# ===============================
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# ===============================
# FLASK APP CONFIG
# ===============================
app = Flask(
    __name__,
    template_folder=resource_path("templates"),
    static_folder=resource_path("static")
)

app.config['SECRET_KEY'] = 'boac_system_secure_key_123'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///users.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

UPLOAD_FOLDER = os.path.join(app.static_folder, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ===============================
# DATABASE MODEL
# ===============================
class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(20), unique=True, nullable=False)
    password = db.Column(db.String(60), nullable=False)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ===============================
# HELPERS
# ===============================
def generate_qr_base64(data):
    """Generates a QR code as a base64 string for a single resident."""
    qr_content = (
        f"MUNICIPALITY OF BOAC ID\n"
        f"----------------------\n"
        f"ID NO: {data.get('id_number')}\n"
        f"NAME: {data.get('full_name')}\n"
        f"POSITION: {data.get('position')}\n"
        f"OFFICE: {data.get('office')}\n"
        f"EMERGENCY: {data.get('contact_name')} ({data.get('contact_number')})"
    )
    qr = qrcode.QRCode(version=None, box_size=10, border=2)
    qr.add_data(qr_content)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="#004422", back_color="white")
    qr_buffer = io.BytesIO()
    qr_img.save(qr_buffer, format="PNG")
    return base64.b64encode(qr_buffer.getvalue()).decode("utf-8")

# ===============================
# EXCEL LOGGING
# ===============================
EXCEL_FILE = "Boac_ID_Database.xlsx"
EXCEL_HEADERS = [
    "Date Generated", "ID Number", "Full Name", "Nickname", 
    "Position", "Office", "Contact Person", "Contact Number", 
    "Address", "Photo Path"
]

def log_to_excel(data):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "ID Records"
        ws.append(EXCEL_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="004422")
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        data["id_number"],
        data["full_name"],
        data["nickname"],
        data["position"],
        data["office"],
        data["contact_name"],
        data["contact_number"],
        data["address"],
        data["photo_filename"]
    ])
    wb.save(EXCEL_FILE)

def get_all_records():
    records = []
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            records.append({
                "date": row[0],
                "id_number": row[1],
                "full_name": row[2],
                "nickname": row[3],
                "position": row[4],
                "office": row[5],
                "contact_name": row[6],
                "contact_number": row[7],
                "address": row[8],
                "photo_filename": row[9]
            })
    return records

# ===============================
# ROUTES
# ===============================

@app.route("/")
@login_required
def index():
    return render_template("index.html", current_user=current_user)

@app.route("/generate", methods=["POST"])
@login_required
def generate_id():
    id_number = request.form.get("id_number", "").strip().upper()
    full_name = request.form.get("full_name", "").strip().upper()
    
    photo_filename = "default.png"
    file = request.files.get("photo_file")
    if file and file.filename:
        timestamp = datetime.now().strftime('%H%M%S')
        photo_filename = f"{id_number}_{timestamp}.png"
        photo_path = os.path.join(UPLOAD_FOLDER, photo_filename)
        file.save(photo_path)

    data = {
        "id_number": id_number,
        "full_name": full_name,
        "nickname": request.form.get("nickname", "").strip().upper(),
        "position": request.form.get("position", "").strip().upper(),
        "office": request.form.get("office", "").strip().upper(),
        "contact_name": request.form.get("contact_name", "").strip().upper(),
        "contact_number": request.form.get("contact_number", "").strip(),
        "address": request.form.get("address", "").strip().upper(),
        "photo_filename": photo_filename
    }

    log_to_excel(data)
    
    flash(f"ID for {full_name} is ready!", "success")
    return redirect(url_for('id_preview', selected_id=id_number, is_new=True))

@app.route("/id-preview")
@login_required
def id_preview():
    records = get_all_records()
    if not records:
        flash("No resident records found.", "info")
        return redirect(url_for('index'))
    
    selected_id = request.args.get('selected_id')
    is_new = request.args.get('is_new', False)
    
    if selected_id:
        data = next((r for r in records if str(r['id_number']) == str(selected_id)), records[-1])
    else:
        data = records[-1]
    
    qr_base64 = generate_qr_base64(data)
    
    return render_template("id_preview.html", 
                           records=records, 
                           selected_record=data, 
                           qr_code=qr_base64,
                           is_new=is_new,
                           current_user=current_user)

@app.route("/batch")
@login_required
def batch_view():
    records = get_all_records()
    # QR codes are generated here so the batch template has them ready for the back side
    for res in records:
        res['qr_base64'] = generate_qr_base64(res)
        
    return render_template("batch_view.html", records=records, current_user=current_user)

# ===============================
# AUTHENTICATION ROUTES
# ===============================

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == "POST":
        user = User.query.filter_by(username=request.form.get("username")).first()
        if user and bcrypt.check_password_hash(user.password, request.form.get("password")):
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Invalid credentials.', 'danger')
    return render_template("login.html")

@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        username = request.form.get("username")
        hashed_pw = bcrypt.generate_password_hash(request.form.get("password")).decode('utf-8')
        if User.query.filter_by(username=username).first():
            flash('Username exists!', 'danger')
            return redirect(url_for('signup'))
        new_user = User(username=username, password=hashed_pw)
        db.session.add(new_user)
        db.session.commit()
        flash('Account created!', 'success')
        return redirect(url_for('login'))
    return render_template("signup.html")

@app.route("/logout")
def logout():
    logout_user()
    return redirect(url_for('login'))

# ===============================
# MAIN EXECUTION
# ===============================
if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    # Setting debug=True enables the auto-reloader
    app.run(host="0.0.0.0", port=5000, debug=True)